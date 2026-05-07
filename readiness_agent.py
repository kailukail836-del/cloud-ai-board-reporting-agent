"""Workbook readiness agent for the V2 Cloud.AI FP&A workflow."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


REQUIRED_SHEETS = [
    "SaaS KPI Dashboard",
    "P&L",
    "Cash Runway",
    "Scenario Analysis",
    "Board Analysis",
    "Board Deck Input",
]

CORE_ANALYSIS_SHEETS = [
    "P&L",
    "Cash Runway",
    "SaaS KPI Dashboard",
    "Scenario Analysis",
]


def _sheet_has_label(ws, label: str) -> bool:
    """Return True when a worksheet contains a label anywhere in its used area."""

    target = label.strip().lower()
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and value.strip().lower() == target:
                return True
    return False


def _sheet_has_enough_content(ws, minimum_cells: int = 8) -> bool:
    """Check whether a sheet has enough populated cells to be useful."""

    populated = 0
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value not in (None, ""):
                populated += 1
                if populated >= minimum_cells:
                    return True
    return False


def inspect_workbook_readiness(workbook_path: Path) -> dict[str, Any]:
    """Inspect workbook structure before the heavier FP&A workflow runs."""

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    available_sheets = list(wb.sheetnames)
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in available_sheets]
    has_core_sheets = {sheet: sheet in available_sheets for sheet in CORE_ANALYSIS_SHEETS}
    missing_items = list(missing_sheets)
    warnings: list[str] = []
    score = 100

    score -= 10 * len(missing_sheets)

    for sheet in CORE_ANALYSIS_SHEETS:
        if sheet not in available_sheets:
            warnings.append(f"Core analysis sheet missing: {sheet}.")

    if "Board Analysis" not in available_sheets:
        warnings.append("Board Analysis sheet is missing; narrative context will be thinner.")
    if "Board Deck Input" not in available_sheets:
        warnings.append("Board Deck Input sheet is missing; deck-specific guidance will be unavailable.")

    structure_checks = {
        "p_and_l_has_line_item_header": False,
        "cash_runway_has_line_item_header": False,
        "kpi_dashboard_has_month_header": False,
        "scenario_analysis_has_scenario_header": False,
    }

    if "P&L" in available_sheets:
        structure_checks["p_and_l_has_line_item_header"] = _sheet_has_label(wb["P&L"], "Line Item")
    if "Cash Runway" in available_sheets:
        structure_checks["cash_runway_has_line_item_header"] = _sheet_has_label(wb["Cash Runway"], "Line Item")
    if "SaaS KPI Dashboard" in available_sheets:
        structure_checks["kpi_dashboard_has_month_header"] = _sheet_has_label(wb["SaaS KPI Dashboard"], "Month")
    if "Scenario Analysis" in available_sheets:
        structure_checks["scenario_analysis_has_scenario_header"] = _sheet_has_label(wb["Scenario Analysis"], "Scenario")

    for check_name, passed in structure_checks.items():
        if not passed:
            score -= 8
            missing_items.append(check_name)
            warnings.append(f"Structure check failed: {check_name}.")

    for sheet in REQUIRED_SHEETS:
        if sheet in available_sheets and not _sheet_has_enough_content(wb[sheet]):
            score -= 5
            warnings.append(f"{sheet} has very little populated content.")

    enough_structure = all(structure_checks.values()) and all(has_core_sheets.values())
    score = max(0, min(100, score))

    if score >= 85 and enough_structure:
        status = "Ready"
    elif score >= 55:
        status = "Needs Review"
    else:
        status = "Not Ready"

    return {
        "readiness_score": score,
        "status": status,
        "available_sheet_names": available_sheets,
        "required_sheets": REQUIRED_SHEETS,
        "missing_sheets": missing_sheets,
        "has_board_analysis": "Board Analysis" in available_sheets,
        "has_board_deck_input": "Board Deck Input" in available_sheets,
        "has_core_analysis_sheets": has_core_sheets,
        "has_enough_structure_for_board_analysis": enough_structure,
        "structure_checks": structure_checks,
        "missing_items": missing_items,
        "manual_review_warnings": warnings,
    }
