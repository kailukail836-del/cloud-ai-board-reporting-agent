"""Financial model audit agent for the V2 Cloud.AI FP&A workflow."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from extract_metrics import TARGET_SHEETS, clean_value


FORMULA_ERROR_MARKERS = {"#NAME?", "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NULL!", "#NUM!"}
MONTHLY_REQUIRED_SHEETS = ["P&L", "Cash Runway"]
KPI_REQUIRED_KEYS = [
    "ending_arr",
    "ending_mrr",
    "ending_customers",
    "gross_margin",
    "ebitda_margin",
    "ltv_cac",
    "funding_need",
]


def _find_header_row(ws, header_label: str) -> int | None:
    """Find a 1-based row number containing a target header label."""

    target = header_label.lower()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == target:
                return cell.row
    return None


def _monthly_headers(ws) -> list[Any]:
    """Return monthly headers after the Line Item column."""

    header_row = _find_header_row(ws, "Line Item")
    if header_row is None:
        return []
    values = [cell.value for cell in ws[header_row]]
    first_used = next((idx for idx, value in enumerate(values) if value not in (None, "")), None)
    if first_used is None:
        return []
    return [clean_value(value) for value in values[first_used + 1 :] if clean_value(value) not in (None, "")]


def _blank_forecast_periods(ws) -> list[str]:
    """Find monthly columns that have a header but no data below it."""

    header_row = _find_header_row(ws, "Line Item")
    if header_row is None:
        return []

    blanks = []
    for col_idx in range(2, ws.max_column + 1):
        header = clean_value(ws.cell(header_row, col_idx).value)
        if header in (None, ""):
            continue

        values = [
            clean_value(ws.cell(row_idx, col_idx).value)
            for row_idx in range(header_row + 1, ws.max_row + 1)
        ]
        if not any(value not in (None, "") for value in values):
            blanks.append(str(header))
    return blanks


def _formula_errors(workbook_path: Path) -> list[dict[str, str]]:
    """Scan formula and cached-value views for common spreadsheet error markers."""

    issues: list[dict[str, str]] = []
    for data_only in (False, True):
        wb = openpyxl.load_workbook(workbook_path, data_only=data_only, read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.strip().upper() in FORMULA_ERROR_MARKERS:
                        issues.append(
                            {
                                "sheet": ws.title,
                                "cell": cell.coordinate,
                                "error": value.strip(),
                                "view": "cached_value" if data_only else "formula",
                            }
                        )
    return issues


def _cash_risk_issues(metrics: dict[str, Any]) -> list[dict[str, Any]]:
    """Return negative ending cash months from extracted Cash Runway metrics."""

    row = (
        metrics.get("sheets", {})
        .get("Cash Runway", {})
        .get("monthly_table", {})
        .get("rows", {})
        .get("ending_cash", {})
    )
    issues = []
    for month, value in row.get("values", {}).items():
        try:
            if value is not None and float(value) < 0:
                issues.append({"month": str(month), "ending_cash": value})
        except (TypeError, ValueError):
            continue
    return issues


def _risky_assumptions(metrics: dict[str, Any]) -> list[str]:
    """Flag high-level SaaS assumptions and outputs that deserve review."""

    summary = metrics.get("summary_metrics", {})
    scenarios = metrics.get("sheets", {}).get("Scenario Analysis", {}).get("scenarios", [])
    risks = []

    if summary.get("ending_cash") is not None and float(summary["ending_cash"]) < 0:
        risks.append("Ending cash is negative in the extracted forecast.")
    if summary.get("funding_need") is not None and float(summary["funding_need"]) > 0:
        risks.append("The model requires funding to maintain the minimum cash balance.")
    if summary.get("ltv_cac") is not None and float(summary["ltv_cac"]) > 8:
        risks.append("LTV:CAC is very high; confirm CAC is fully loaded and churn assumptions are realistic.")
    if summary.get("gross_margin") is not None and float(summary["gross_margin"]) < 0.65:
        risks.append("Gross margin is below a typical SaaS target range.")

    for scenario in scenarios:
        churn = scenario.get("Monthly Churn")
        growth = scenario.get("New Customer Growth")
        if churn is not None and float(churn) >= 0.04:
            risks.append(f"{scenario.get('Scenario', 'A scenario')} uses monthly churn at or above 4%.")
        if growth is not None and float(growth) >= 0.08:
            risks.append(f"{scenario.get('Scenario', 'A scenario')} assumes aggressive monthly new-customer growth.")

    return risks


def audit_financial_model(workbook_path: Path, metrics: dict[str, Any]) -> dict[str, Any]:
    """Run a basic model health check and return scoring plus fix guidance."""

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    available_sheets = list(wb.sheetnames)
    missing_sheets = [sheet for sheet in TARGET_SHEETS if sheet not in available_sheets]
    issue_list: list[dict[str, Any]] = []
    recommended_fixes: list[str] = []
    score = 100

    if missing_sheets:
        score -= 10 * len(missing_sheets)
        issue_list.append({"severity": "High", "issue": "Missing required sheets", "detail": missing_sheets})
        recommended_fixes.append("Add or rename required reporting tabs so they match the expected V2 workflow names.")

    monthly_header_checks = {}
    blank_periods = {}
    for sheet in MONTHLY_REQUIRED_SHEETS:
        if sheet not in available_sheets:
            continue
        headers = _monthly_headers(wb[sheet])
        blanks = _blank_forecast_periods(wb[sheet])
        monthly_header_checks[sheet] = headers
        blank_periods[sheet] = blanks
        if not headers:
            score -= 12
            issue_list.append({"severity": "High", "issue": "Missing monthly headers", "sheet": sheet})
            recommended_fixes.append(f"Add month headers to {sheet} after the Line Item column.")
        if blanks:
            score -= min(10, 2 * len(blanks))
            issue_list.append({"severity": "Medium", "issue": "Blank forecast periods", "sheet": sheet, "periods": blanks})
            recommended_fixes.append(f"Populate blank forecast months in {sheet}: {', '.join(blanks)}.")

    formula_errors = _formula_errors(workbook_path)
    if formula_errors:
        score -= min(30, 5 * len(formula_errors))
        issue_list.append({"severity": "High", "issue": "Formula error indicators found", "detail": formula_errors[:20]})
        recommended_fixes.append("Resolve spreadsheet formula errors before using the pack with stakeholders.")

    negative_cash_months = _cash_risk_issues(metrics)
    if negative_cash_months:
        score -= min(15, 3 * len(negative_cash_months))
        issue_list.append({"severity": "High", "issue": "Negative cash months", "detail": negative_cash_months})
        recommended_fixes.append("Add financing, delay spend, or revise plan timing before cash turns negative.")

    scenarios = metrics.get("sheets", {}).get("Scenario Analysis", {}).get("scenarios", [])
    if len(scenarios) < 3:
        score -= 8
        issue_list.append({"severity": "Medium", "issue": "Missing scenario rows", "detail": f"Found {len(scenarios)} rows."})
        recommended_fixes.append("Include at least conservative, base, and optimistic scenarios.")

    missing_kpis = [key for key in KPI_REQUIRED_KEYS if metrics.get("summary_metrics", {}).get(key) is None]
    if missing_kpis:
        score -= 4 * len(missing_kpis)
        issue_list.append({"severity": "Medium", "issue": "Missing KPI values", "detail": missing_kpis})
        recommended_fixes.append("Confirm dashboard KPI labels and values are present and numeric.")

    risky_assumptions = _risky_assumptions(metrics)
    if risky_assumptions:
        score -= min(12, 3 * len(risky_assumptions))
        issue_list.append({"severity": "Medium", "issue": "Risky assumptions or outputs", "detail": risky_assumptions})
        recommended_fixes.append("Flag risky assumptions in the board narrative and show downside sensitivity.")

    score = max(0, min(100, score))
    if score >= 85:
        status = "Green"
    elif score >= 60:
        status = "Amber"
    else:
        status = "Red"

    return {
        "model_health_score": score,
        "status": status,
        "issue_list": issue_list,
        "recommended_fixes": recommended_fixes,
        "checks": {
            "missing_required_sheets": missing_sheets,
            "monthly_headers": monthly_header_checks,
            "blank_forecast_periods": blank_periods,
            "formula_errors": formula_errors,
            "negative_cash_months": negative_cash_months,
            "missing_scenario_rows": len(scenarios) < 3,
            "missing_kpi_values": missing_kpis,
            "risky_assumptions": risky_assumptions,
        },
    }
